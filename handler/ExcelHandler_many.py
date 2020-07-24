# coding=gbk

# pip install openpyxl
import os
from openpyxl import load_workbook
from configuration.Path import path as pth
from handler.LogHandler import log
from handler.ConfigHandler import config


class ExcelHandler:
    """������ͬʱ�������sheetҳ"""

    def __init__(self, filename):
        self.filename = filename

    def open(self, sheet_name):
        """���ع�����"""
        try:
            self.wb = load_workbook(self.filename)
            if isinstance(sheet_name, str):
                self.sheet = self.wb[sheet_name]
            else:
                self.sheet = self.wb.worksheets[sheet_name]
        except Exception as e:
            log.error(e)
            raise e

    def get_headers(self, sheet):
        """��ȡsheetҳ����ͷ"""
        self.open(sheet)
        headers = [c.value for c in self.sheet[1]]
        self.wb.close()
        return headers

    def read(self, sheet, end_col=5):
        """��ȡ���ݣ������б�Ƕ���ֵ�"""
        self.open(sheet)
        max_row = self.sheet.max_row
        data = []
        for row in range(2, max_row+1):         # �ڶ��п�ʼ��ȡ
            row_data = {}
            for col in range(1, end_col+1):
                row_data[self.sheet.cell(1, col).value] = self.sheet.cell(row, col).value
            data.append(row_data)
        self.wb.close()
        return data

    def read_to_list(self, sheet, end_col=5):
        """��ȡ���ݣ�����Ƕ���б�"""
        self.open(sheet)
        rows_data = list(self.sheet.rows)[1:]       # ��ȡ���������ݣ���������һ��
        data = []
        for row in rows_data:
            row_data = []
            for cell in row[:end_col]:
                row_data.append(cell.value)
            data.append(row_data)
        self.wb.close()
        return data

    def write(self, sheet, row, col, value):
        """д������"""
        self.open(sheet)
        self.sheet.cell(row, col).value = value
        self.save()

    def save(self):
        """���沢�ر�"""
        self.wb.save(self.filename)
        self.wb.close()


file_name = config.get('excel', 'filename')
file_path = os.path.join(pth.data_path, file_name)

ex = ExcelHandler(filename=file_path)

if __name__ == '__main__':
    print(ex.write('login', 2, 6, 'xx'))
