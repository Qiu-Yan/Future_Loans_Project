# coding=gbk

import os
# pip install openpyxl
from openpyxl import load_workbook
from handler.ConfigHandler import config
from handler.LogHandler import log
from configuration.Path import path as pth


class ExcelHandler:
    """���õ���sheet�����ݴ���"""

    def __init__(self, filename, sheet_name):
        """filename�� excel�ļ�·���� sheet�� ������"""
        self.filename = filename
        self.sheet_name = sheet_name

    def open(self):
        """���ع���������ȡ��"""
        try:
            self.wb = load_workbook(filename=self.filename)
            if isinstance(self.sheet_name, str):
                self.sheet = self.wb[self.sheet_name]
            else:
                self.sheet = self.wb.worksheets[self.sheet_name]
        except Exception as e:
            log.error(e)
            raise e

    def get_headers(self):
        """��ȡ��ͷ"""
        self.open()
        header = [cell.value for cell in self.sheet[1]]
        self.wb.close()
        return header

    def read(self, end_col=5):
        """��ȡ���ݣ������б�Ƕ���ֵ�, end_col, ������"""
        self.open()
        max_row = self.sheet.max_row
        data = []
        for row in range(2, max_row+1):
            row_data = {}
            for col in range(1, end_col+1):
                row_data[self.sheet.cell(1, col).value] = self.sheet.cell(row, col).value
            data.append(row_data)
        self.wb.close()
        return data

    def read_to_list(self, end_col=5):
        """��ȡ�������ݣ�����Ƕ���б�"""
        self.open()
        rows_data = list(self.sheet.rows)[1:]
        data = []
        for row in rows_data:
            row_data = []
            for cell in row[:end_col]:
                row_data.append(cell.value)
            data.append(row_data)
        self.wb.close()
        return data

    def write(self, row, col, value):
        """д������"""
        self.open()
        self.sheet.cell(row, col).value = value
        self.save()

    def save(self):
        """���沢�ر�"""
        self.wb.save(self.filename)
        self.wb.close()


"""��ȡ�����ļ���excel�ļ�����"""
file_name = config.get('excel', 'filename')
file_path = os.path.join(pth.data_path, file_name)


def load_sheet(sheet):
    return ExcelHandler(file_path, sheet)


if __name__ == '__main__':
    # file_name = config.get('excel', 'filename')
    # file_path = os.path.join(pth.data_path, file_name)
    # sheet_name = config.get('excel', 'register_sheet')
    ex = load_sheet('register')
    print(ex.get_headers())
